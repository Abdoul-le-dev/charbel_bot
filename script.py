import re
import asyncio
import sqlite3
import traceback
import unicodedata
from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardRemove
from telegram.ext import (
    ChatJoinRequestHandler, CallbackQueryHandler, Application,
    CommandHandler, MessageHandler, filters, ContextTypes, ConversationHandler
)
from database.database import init_db, upsert_user, log_member, get_file_id, save_file_id

# ── Module sondage (totalement autonome) ─────────────────────────────────────
from sondage import init_sondage_db, register_sondage_handlers

TOKEN = "8609131464:AAGK5k1jkLJvY1OSvHcR3YPnwqEqOFeWuAs"

ADMIN_IDS      = {6992809421, 6799962131}
ADMIN_USERNAME = "@Faiseur2Rois"

# ── Lien vidéos débutants ─────────────────────────────────────────────────────
LIEN_YOUTUBE_DEBUTANTS = "https://www.youtube.com/live/Cr-uCj8iQX4?si=L1fFk_8zl3dnW8ii"

PLACES_RESTANTES = 47
PLACES_TOTALES   = 150

# ── États inscription ─────────────────────────────────────────────────────────
(PRENOM, PRENOM_CONFIRM, WHATSAPP, PAYS, DEJA_TRADE,
 INTERET, PRESENCE, FREIN, CONFIRMATION) = range(9)

# ── États broadcast ───────────────────────────────────────────────────────────
BC_CIBLE, BC_FORMAT, BC_MEDIA, BC_TEXT = range(9, 13)

# ── États création de catégorie ───────────────────────────────────────────────
CAT_NOM = 13


# ════════════════════════════════════════════════════════════════════════════
# ── BASE DE DONNÉES ───────────────────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

def db():
    conn = sqlite3.connect("preinscriptions.db")
    conn.row_factory = sqlite3.Row
    return conn


def _migrate_db():
    print('start')
    with db() as conn:
        cur = conn.cursor()
        colonnes_a_ajouter = {
            "categorie":  f"TEXT DEFAULT '{EVENEMENT_ACTUEL}'",
            "last_seen":  "DATETIME",
            # ── Nouvelles colonnes formulaire Trading Pour Tous ──
            "pays":       "TEXT",
            "deja_trade": "TEXT",
            "interet":    "TEXT",
            "presence":   "TEXT",
            "frein":      "TEXT",
        }
        cur.execute("PRAGMA table_info(users)")
        existantes = {row["name"] for row in cur.fetchall()}
        for col, definition in colonnes_a_ajouter.items():
            if col not in existantes:
                cur.execute(f"ALTER TABLE users ADD COLUMN {col} {definition}")

        cur.execute("""
            CREATE TABLE IF NOT EXISTS categories (
                id        INTEGER PRIMARY KEY AUTOINCREMENT,
                nom       TEXT UNIQUE NOT NULL,
                creee_le  DATETIME DEFAULT (datetime('now')),
                active    INTEGER DEFAULT 1
            )
        """)
        cur.execute(
            "INSERT OR IGNORE INTO categories (nom) VALUES (?)",
            (EVENEMENT_ACTUEL,)
        )
        cur.execute("""
            CREATE TABLE IF NOT EXISTS messages_libres (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                telegram_id INTEGER,
                texte       TEXT,
                recu_le     DATETIME DEFAULT (datetime('now'))
            )
        """)
        conn.commit()

        print("Migration terminée. Colonnes ajoutées et tables créées si nécessaire.")


def _get_last_category() -> str | None:
    with db() as conn:
        row = conn.execute(
            "SELECT nom FROM categories WHERE active=1 ORDER BY creee_le DESC LIMIT 1"
        ).fetchone()
    return row["nom"] if row else None


EVENEMENT_ACTUEL = _get_last_category()


def _is_already_registered(user_id: int) -> bool:
    """Vérifie l'inscription selon les NOUVEAUX champs obligatoires."""
    with db() as conn:
        row = conn.execute(
            """SELECT 1 FROM users
               WHERE telegram_id = ? AND completed = 1
                 AND prenom IS NOT NULL AND whatsapp IS NOT NULL
                 AND pays IS NOT NULL AND deja_trade IS NOT NULL
                 AND interet IS NOT NULL AND presence IS NOT NULL
                 AND frein IS NOT NULL""",
            (user_id,)
        ).fetchone()
    return row is not None


def _get_incomplete_users() -> list[dict]:
    with db() as conn:
        rows = conn.execute("""
            SELECT m.telegram_id, u.prenom
            FROM members_log m
            LEFT JOIN users u ON u.telegram_id = m.telegram_id
            WHERE u.telegram_id IS NULL
               OR u.completed = 0
            ORDER BY m.joined_at DESC
        """).fetchall()
    return [dict(r) for r in rows]


def _get_all_user_ids() -> list[int]:
    with db() as conn:
        rows = conn.execute("""
            SELECT telegram_id FROM users
            WHERE telegram_id IS NOT NULL
            UNION
            SELECT telegram_id FROM members_log
        """).fetchall()
    return [r["telegram_id"] for r in rows]


def _get_users_by_filter(filter_type: str, filter_value: str) -> list[int]:
    """
    Retourne les user_ids selon un filtre.
    filter_type : 'presence' | 'frein' | 'deja_trade' | 'interet' | 'all' | 'incomplets'
    filter_value : la valeur exacte à matcher (ignoré pour 'all' et 'incomplets')
    """
    with db() as conn:
        if filter_type == "all":
            rows = conn.execute("""
                SELECT telegram_id FROM users WHERE telegram_id IS NOT NULL
                UNION
                SELECT telegram_id FROM members_log
            """).fetchall()
        elif filter_type == "incomplets":
            rows = conn.execute("""
                SELECT m.telegram_id
                FROM members_log m
                LEFT JOIN users u ON u.telegram_id = m.telegram_id
                WHERE u.telegram_id IS NULL OR u.completed = 0
            """).fetchall()
        elif filter_type == "complets":
            rows = conn.execute(
                "SELECT telegram_id FROM users WHERE completed = 1"
            ).fetchall()
        else:
            # presence, frein, deja_trade, interet
            rows = conn.execute(
                f"SELECT telegram_id FROM users WHERE {filter_type} = ? AND completed = 1",
                (filter_value,)
            ).fetchall()
    return [r["telegram_id"] for r in rows]


def _get_categories() -> list[str]:
    with db() as conn:
        rows = conn.execute(
            "SELECT nom FROM categories WHERE active = 1 ORDER BY id"
        ).fetchall()
    return [r["nom"] for r in rows]


def _ajouter_categorie(nom: str):
    with db() as conn:
        conn.execute("INSERT OR IGNORE INTO categories (nom) VALUES (?)", (nom,))
        conn.commit()


def _touch_last_seen(user_id: int):
    with db() as conn:
        conn.execute(
            "UPDATE users SET last_seen = datetime('now') WHERE telegram_id = ?",
            (user_id,)
        )
        conn.commit()


def _log_message(user_id: int, texte: str):
    with db() as conn:
        conn.execute(
            "INSERT INTO messages_libres (telegram_id, texte) VALUES (?, ?)",
            (user_id, texte)
        )
        conn.commit()


# ════════════════════════════════════════════════════════════════════════════
# ── NETTOYAGE DU PRÉNOM ───────────────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

def _nettoyer_prenom(texte: str) -> str:
    t = texte.strip()
    t = t.replace("\u2019", "'").replace("\u2018", "'")
    t = t.lower()
    t = "".join(c for c in t if unicodedata.category(c) not in ("So", "Sm", "Sk", "Cn"))

    for ancien, nouveau in [
        ("mappelle", " "), ("cest", " "), ("mest", " "),
        ("jsuis", " "),    ("chuis", " "),
    ]:
        t = t.replace(ancien, nouveau)

    parasites = [
        "tout le monde m'appelle", "vous pouvez m'appeler", "tu peux m'appeler",
        "ils m'appellent", "on m'appelle",
        "je me présente", "je me presente",
        "ravi de te rencontrer", "ravi de vous rencontrer",
        "comment ça va", "comment ca va",
        "bonne journée", "bonne soirée", "bonne nuit",
        "bien sûr", "bien sur",
        "je m'appelle", "j'me appelle",
        "mon prénom c'est", "mon prenom c'est",
        "mon prénom est", "mon prenom est",
        "mon nom c'est", "mon nom est",
        "je me nomme", "je me nome",
        "moi c'est", "c'est moi",
        "my name is", "my name's",
        "they call me", "call me",
        "people call me", "everyone calls me", "you can call me",
        "je suis", "je sui",
        "i am", "i'm",
        "avec plaisir",
        "bonjour", "bonsoir", "salut", "coucou", "hello",
        "hey", "wesh", "yo", "slt", "salu", "hi",
        "enchanté", "enchante", "enchantée",
        "voilà", "voila", "voici",
        "exactement", "exact",
        "ouais", "ouas", "oui",
        "alors", "donc", "bien",
        "prénom", "prenom", "appelle", "appeler", "nomme",
        "ça va", "ca va",
        "ok",
        "mon", "ma", "mes", "nom",
        "moi", "moa", "mwa",
        "suis", "est", "dj", "je", "et",
    ]

    for p in parasites:
        pattern = re.escape(p)
        t = re.sub(r"(?<![a-zà-ÿ])" + pattern + r"(?![a-zà-ÿ])", " ", t)

    t = re.sub(r"[^\w\s\u00C0-\u00FF\-']", " ", t)
    t = re.sub(r"(?<![a-zà-ÿ])'|'(?![a-zà-ÿ])", " ", t)
    t = re.sub(r"\s+", " ", t).strip()

    def cap_mot(m: str) -> str:
        return "-".join(p.capitalize() for p in m.split("-"))

    mots = [cap_mot(m) for m in t.split() if len(m) > 1]
    return " ".join(mots) if mots else texte.strip().split()[-1].capitalize()


def _extraire_prenom(texte: str) -> tuple[str, bool]:
    brut      = texte.strip()
    mots_brut = brut.split()

    if len(mots_brut) <= 2 and len(brut) <= 20:
        return brut.title(), False

    nettoye = _nettoyer_prenom(brut)
    return nettoye, True


# ════════════════════════════════════════════════════════════════════════════
# ── KEYBOARDS ────────────────────────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

def kb_prenom_confirm(prenom: str):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(f"✅ Oui, c'est bien {prenom}", callback_data="prenom_oui")],
        [InlineKeyboardButton("✏️ Non, je vais corriger",     callback_data="prenom_non")],
    ])

def kb_deja_trade():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Oui", callback_data="trade_oui")],
        [InlineKeyboardButton("❌ Non", callback_data="trade_non")],
    ])

def kb_interet():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("💰 Un revenu complémentaire",       callback_data="int_revenu")],
        [InlineKeyboardButton("📚 Apprendre une nouvelle compétence", callback_data="int_competence")],
        [InlineKeyboardButton("🕊️ L'indépendance financière",       callback_data="int_independance")],
        [InlineKeyboardButton("✍️ Autre",                            callback_data="int_autre")],
    ])

def kb_presence():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Oui, les deux jours",   callback_data="pres_deux")],
        [InlineKeyboardButton("⚠️ Un seul des deux jours", callback_data="pres_un")],
    ])

def kb_frein():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("⏱️ Le manque de temps",           callback_data="frein_temps")],
        [InlineKeyboardButton("😰 La peur de perdre de l'argent", callback_data="frein_peur")],
        [InlineKeyboardButton("🤷 Je ne sais pas par où commencer", callback_data="frein_commencer")],
        [InlineKeyboardButton("✍️ Autre",                          callback_data="frein_autre")],
    ])

def kb_confirmation():
    return InlineKeyboardMarkup([[
        InlineKeyboardButton("🚀 Confirmer mon inscription !", callback_data="confirme")
    ]])

def kb_relance():
    return InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ Je finalise mon inscription", callback_data="relance_go")
    ]])


# ════════════════════════════════════════════════════════════════════════════
# ── VIDÉO DE BIENVENUE ───────────────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

async def send_welcome_video(bot, user_id: int):
    log_member(user_id)
    upsert_user(user_id, categorie=EVENEMENT_ACTUEL)

    video_name = "welcomes_2"
    file_id    = get_file_id(video_name)
    caption    = (
        "🚀 *Bienvenue dans Trading Pour Tous !*\n\n"
        "Tu es sur le point de réserver ta place à la formation gratuite :\n"
        "*« Trading Pour Tous »*\n\n"
        "📅 *2 et 3 septembre à 21h00* (heure du Bénin)\n"
        "🎥 En direct uniquement — places limitées"
    )

    if file_id:
        await bot.send_video(chat_id=user_id, video=file_id,
                             caption=caption, parse_mode="Markdown")
    else:
        msg = await bot.send_video(
            chat_id=user_id,
            video=open("video/welcome.mp4", "rb"),
            caption=caption, parse_mode="Markdown"
        )
        save_file_id(video_name, msg.video.file_id)

    await bot.send_message(
        chat_id=user_id,
        text=(
            "⚠️ *Il ne reste que peu de places*\n"
            "Les places s'envolent vite. Sécurise la tienne maintenant.\n\n"
            "👇 Clique ici pour confirmer ta place :\n\n"
            "/JeMEnregistre"
        ),
        parse_mode="Markdown"
    )


async def _send_welcome_safe(bot, user_id: int):
    try:
        await send_welcome_video(bot, user_id)
    except Exception as e:
        print(f"Erreur bienvenue uid={user_id} : {e}")


async def _reply_already_registered(bot, user_id: int):
    await bot.send_message(
        chat_id=user_id,
        text=(
            "✅ *Tu es déjà inscrit à la formation Trading Pour Tous*\n\n"
            "Pas besoin de t'enregistrer une deuxième fois 😊\n\n"
            "Rendez-vous les *2 et 3 septembre à 21h00* (heure du Bénin).\n\n"
            "Tu recevras le lien du live et tous les rappels "
            "par *WhatsApp et Telegram* avant l'événement.\n\n"
            "Hâte de te voir en ligne 🔥"
        ),
        parse_mode="Markdown"
    )


# ════════════════════════════════════════════════════════════════════════════
# ── HANDLERS ENTRÉE ──────────────────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

async def _handle_join(bot, user_id: int):
    try:
        await send_welcome_video(bot, user_id)
    except Exception as e:
        print(f"Erreur join uid={user_id} : {e}")


async def approve_join_request(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.chat_join_request.from_user.id
    try:
        await update.chat_join_request.approve()
    except Exception as e:
        print(f"approve() uid={user_id} : {e}")

    asyncio.create_task(_handle_join(context.bot, user_id))


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    args    = context.args

    if args and args[0] == "JeMenregistre":
        await update.message.reply_text(
            "Super 🎉 Clique sur /JeMEnregistre pour t'inscrire à la formation.",
            parse_mode="Markdown"
        )
        return

    _touch_last_seen(user_id)
    if _is_already_registered(user_id):
        await _reply_already_registered(context.bot, user_id)
        return
    asyncio.create_task(_send_welcome_safe(context.bot, user_id))


# ════════════════════════════════════════════════════════════════════════════
# ── MESSAGES LIBRES → REDIRECTION ADMIN ──────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

async def message_libre(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message or not update.message.text:
        return

    user    = update.effective_user
    user_id = user.id
    texte   = update.message.text.lower().strip()

    _touch_last_seen(user_id)
    _log_message(user_id, texte)

    if "present" in texte or "présent" in texte:
        reponse = (
            "Je suis très content de savoir que tu seras là !\n\n"
            "N'oublie pas, c'est les 2 et 3 septembre à 21h heure de Cotonou.\n\n"
            "Je t'enverrai le lien du live juste avant, ici et sur WhatsApp si possible."
        )
        await update.message.reply_text(reponse)
    elif "merci" in texte:
        await update.message.reply_text("jtp !")
    elif "ok" in texte:
        await update.message.reply_text("Super !")
    else:
        await update.message.reply_text(
            "Ton message a bien été reçu.\n\n"
            f"Pour une réponse rapide, contacte directement "
            f"Charbel sur Telegram : {ADMIN_USERNAME}\n\n"
            "Il te répondra dès que possible."
        )

    username      = f"@{user.username}" if user.username else f"id:{user_id}"
    prenom_tg     = user.first_name or ""
    texte_affiche = update.message.text[:200]

    notif = (
        f"Nouveau message reçu\n\n"
        f"Utilisateur : {prenom_tg} ({username})\n"
        f"ID : {user_id}\n\n"
        f"Message : {texte_affiche}"
    )
    for admin_id in ADMIN_IDS:
        try:
            await context.bot.send_message(chat_id=admin_id, text=notif)
        except Exception as e:
            print(f"Notif admin {admin_id} : {e}")


# ════════════════════════════════════════════════════════════════════════════
# ── RELANCE INSCRIPTIONS INCOMPLÈTES ─────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

async def relancer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Commande réservée à l'administrateur.")
        return

    users = _get_incomplete_users()
    if not users:
        await update.message.reply_text("✅ Aucune inscription incomplète à relancer.")
        return

    await update.message.reply_text(
        f"📤 Relance en cours pour *{len(users)}* utilisateurs...",
        parse_mode="Markdown"
    )
    asyncio.create_task(_broadcast_relance(context.bot, update.effective_user.id, users))


async def _broadcast_relance(bot, admin_id: int, users: list[dict]):
    sent = blocked = erreurs = 0

    for u in users:
        uid    = u["telegram_id"]
        prenom = u["prenom"] or "Hello l'ami"
        try:
            await bot.send_message(
                chat_id=uid,
                text=(
                    f"⚠️ *{prenom}, ton inscription n'a pas encore été validée*\n\n"
                    "Tu as commencé à t'inscrire à la formation gratuite "
                    "*Trading Pour Tous* (2 et 3 septembre à 21h), "
                    "mais tu n'as pas finalisé ta demande.\n\n"
                    "Il ne reste que peu de places et elles partent vite.\n\n"
                    "Clique sur le bouton ci-dessous pour sécuriser ta place :"
                ),
                parse_mode="Markdown",
                reply_markup=kb_relance()
            )
            sent += 1
        except Exception as e:
            err = str(e)
            if "Forbidden" in err or "blocked" in err.lower():
                blocked += 1
            elif "can't initiate" in err:
                erreurs += 1
            else:
                erreurs += 1
                print(f"Relance uid={uid} : {e}")

        await asyncio.sleep(0.1)

    await bot.send_message(
        admin_id,
        f"Relance terminée\n\n"
        f"*{sent}* envoyés\n"
        f"*{blocked}* ont bloqué le bot\n"
        f"*{erreurs}* autres erreurs\n"
        f"Total ciblé : *{len(users)}*",
        parse_mode="Markdown"
    )


async def relance_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    texte_reponse = "Super 🎉 Clique sur /JeMEnregistre pour finaliser ton inscription."
    if query.message:
        await query.message.reply_text(texte_reponse, parse_mode="Markdown")
    else:
        await context.bot.send_message(
            chat_id=query.from_user.id,
            text=texte_reponse,
            parse_mode="Markdown"
        )


# ════════════════════════════════════════════════════════════════════════════
# ── GESTION DES CATÉGORIES ────────────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

async def nouvelle_categorie_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Commande réservée à l'administrateur.")
        return ConversationHandler.END

    cats  = _get_categories()
    liste = "\n".join(f"• {c}" for c in cats) if cats else "_(aucune pour l'instant)_"
    await update.message.reply_text(
        f"*Catégories actuelles :*\n\n{liste}\n\n"
        "Envoie le *nom* du nouvel événement à créer :",
        parse_mode="Markdown"
    )
    return CAT_NOM


async def nouvelle_categorie_nom(update: Update, context: ContextTypes.DEFAULT_TYPE):
    nom = update.message.text.strip()
    if len(nom) < 3:
        await update.message.reply_text("❌ Nom trop court (min 3 caractères). Réessaie :")
        return CAT_NOM
    _ajouter_categorie(nom)
    await update.message.reply_text(
        f"✅ Catégorie *{nom}* créée avec succès.",
        parse_mode="Markdown"
    )
    return ConversationHandler.END


async def nouvelle_categorie_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("❌ Création annulée.")
    return ConversationHandler.END


# ════════════════════════════════════════════════════════════════════════════
# ── BROADCAST AVEC CIBLAGE ───────────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

# Map cible → (filter_type, filter_value, libellé humain)
CIBLES = {
    "1":  ("all",         None,                                       "Tous les utilisateurs"),
    "2":  ("complets",    None,                                       "Inscrits complets uniquement"),
    "3":  ("incomplets",  None,                                       "Inscriptions incomplètes"),
    "4":  ("presence",    "Oui, les deux jours",                      "Présents les 2 jours"),
    "5":  ("presence",    "Un seul des deux jours",                   "Présents un seul jour"),
    "6":  ("frein",       "Le manque de temps",                       "Frein : manque de temps"),
    "7":  ("frein",       "La peur de perdre de l'argent",            "Frein : peur de perdre"),
    "8":  ("frein",       "Je ne sais pas par où commencer",          "Frein : ne sait pas commencer"),
    "9":  ("frein",       "Autre",                                    "Frein : autre"),
    "10": ("deja_trade",  "Oui",                                      "Ont déjà tradé"),
    "11": ("deja_trade",  "Non",                                      "N'ont jamais tradé (débutants)"),
    "12": ("interet",     "Un revenu complémentaire",                 "Intérêt : revenu complémentaire"),
    "13": ("interet",     "Apprendre une nouvelle compétence",        "Intérêt : nouvelle compétence"),
    "14": ("interet",     "L'indépendance financière",                "Intérêt : indépendance"),
    "15": ("interet",     "Autre",                                    "Intérêt : autre"),
}


async def _broadcast_targeted(bot, admin_id: int, data: dict):
    filter_type  = data["filter_type"]
    filter_value = data["filter_value"]
    libelle      = data["libelle"]
    fmt          = data.get("format")
    texte        = data.get("text_content", "")
    media_id     = data.get("media_file_id")

    user_ids = _get_users_by_filter(filter_type, filter_value)
    total    = len(user_ids)

    if total == 0:
        await bot.send_message(admin_id, f"❌ Aucun utilisateur dans la cible : *{libelle}*.", parse_mode="Markdown")
        return
    if fmt in {"2", "3", "4", "5"} and not media_id:
        await bot.send_message(admin_id, "❌ Fichier média manquant. Diffusion annulée.")
        return

    est = round(total * 0.1 / 60, 2)
    await bot.send_message(admin_id,
        f"📤 Envoi en cours\nCible : *{libelle}*\nDestinataires : *{total}*\nEstimé : {est} min",
        parse_mode="Markdown")

    sent = 0
    for idx, uid in enumerate(user_ids, start=1):
        try:
            if fmt == "1":   await bot.send_message(chat_id=uid, text=texte)
            elif fmt == "2": await bot.send_photo(chat_id=uid, photo=media_id, caption=texte)
            elif fmt == "3": await bot.send_video(chat_id=uid, video=media_id, caption=texte)
            elif fmt == "4": await bot.send_photo(chat_id=uid, photo=media_id)
            elif fmt == "5": await bot.send_video(chat_id=uid, video=media_id)
            sent += 1
        except Exception as e:
            print(f"Broadcast uid={uid} : {e}")

        if   total >= 3 and idx == total // 3:
            await bot.send_message(admin_id, "1/3 des messages envoyés")
        elif total >= 3 and idx == (2*total) // 3:
            await bot.send_message(admin_id, "2/3 des messages envoyés")
        elif idx == total:
            await bot.send_message(admin_id,
                f"Diffusion terminée — *{sent}/{total}* messages envoyés à *{libelle}*",
                parse_mode="Markdown")

        await asyncio.sleep(0.1)


async def bc_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Commande réservée à l'administrateur.")
        return ConversationHandler.END

    lignes = ["*Choisis la cible du message :*\n"]
    for num, (_, _, libelle) in CIBLES.items():
        # Affiche aussi le nombre approximatif
        ftype, fval, _ = CIBLES[num]
        try:
            nb = len(_get_users_by_filter(ftype, fval))
        except Exception:
            nb = "?"
        lignes.append(f"*{num}* — {libelle} _({nb})_")

    lignes.append("\n_Réponds avec le numéro de la cible (ex : 4)_")
    await update.message.reply_text(
        "\n".join(lignes),
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove()
    )
    return BC_CIBLE


async def bc_get_cible(update: Update, context: ContextTypes.DEFAULT_TYPE):
    choix = update.message.text.strip()
    if choix not in CIBLES:
        await update.message.reply_text(
            f"❌ Numéro invalide. Réponds avec un nombre entre 1 et {len(CIBLES)}."
        )
        return BC_CIBLE

    ftype, fval, libelle = CIBLES[choix]
    context.user_data["bc_filter_type"]  = ftype
    context.user_data["bc_filter_value"] = fval
    context.user_data["bc_libelle"]      = libelle

    await update.message.reply_text(
        f"✅ Cible sélectionnée : *{libelle}*\n\n"
        "*Format du message à diffuser :*\n\n"
        "1 — Texte seul\n"
        "2 — Image + texte\n"
        "3 — Vidéo + texte\n"
        "4 — Image seule\n"
        "5 — Vidéo seule\n\n"
        "_(max 4096 caractères)_",
        parse_mode="Markdown"
    )
    return BC_FORMAT


async def bc_get_format(update: Update, context: ContextTypes.DEFAULT_TYPE):
    choix = update.message.text.strip()[0]
    if choix not in {"1", "2", "3", "4", "5"}:
        await update.message.reply_text("❌ Chiffre entre 1 et 5 uniquement.")
        return BC_FORMAT
    context.user_data["bc_format"] = choix
    if choix in {"2", "3"}:
        await update.message.reply_text(f"Envoie ton fichier {'image' if choix=='2' else 'vidéo'}.")
        return BC_MEDIA
    await update.message.reply_text(
        "Envoie maintenant ton texte." if choix == "1" else "Envoie ton fichier."
    )
    return BC_TEXT


async def bc_get_media(update: Update, context: ContextTypes.DEFAULT_TYPE):
    choix = context.user_data["bc_format"]
    if choix == "2":
        if not update.message.photo:
            await update.message.reply_text("❌ Ce n'est pas une image. Réessaie.")
            return BC_MEDIA
        context.user_data["bc_media_id"] = update.message.photo[-1].file_id
    elif choix == "3":
        if not update.message.video:
            await update.message.reply_text("❌ Ce n'est pas une vidéo. Réessaie.")
            return BC_MEDIA
        context.user_data["bc_media_id"] = update.message.video.file_id
    await update.message.reply_text("Envoie maintenant le texte associé.")
    return BC_TEXT


async def bc_get_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    admin_id = update.effective_user.id
    choix    = context.user_data["bc_format"]
    if choix == "4":
        if not update.message.photo:
            await update.message.reply_text("❌ Ce n'est pas une image.")
            return BC_TEXT
        context.user_data["bc_media_id"] = update.message.photo[-1].file_id
        context.user_data["bc_text"]     = ""
    elif choix == "5":
        if not update.message.video:
            await update.message.reply_text("❌ Ce n'est pas une vidéo.")
            return BC_TEXT
        context.user_data["bc_media_id"] = update.message.video.file_id
        context.user_data["bc_text"]     = ""
    else:
        if not update.message.text:
            await update.message.reply_text("❌ Merci d'envoyer du texte.")
            return BC_TEXT
        context.user_data["bc_text"] = update.message.text

    await update.message.reply_text("Diffusion lancée en arrière-plan...")
    asyncio.create_task(_broadcast_targeted(context.bot, admin_id, {
        "filter_type":   context.user_data["bc_filter_type"],
        "filter_value":  context.user_data["bc_filter_value"],
        "libelle":       context.user_data["bc_libelle"],
        "format":        choix,
        "text_content":  context.user_data.get("bc_text", ""),
        "media_file_id": context.user_data.get("bc_media_id"),
    }))
    return ConversationHandler.END


async def bc_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("❌ Diffusion annulée.")
    return ConversationHandler.END


# ════════════════════════════════════════════════════════════════════════════
# ── /stats ────────────────────────────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

async def stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Commande réservée à l'administrateur.")
        return

    with db() as conn:
        total    = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        complets = conn.execute("SELECT COUNT(*) FROM users WHERE completed=1").fetchone()[0]
        try:
            members = conn.execute("SELECT COUNT(*) FROM members_log").fetchone()[0]
        except Exception:
            members = "—"
        try:
            msgs = conn.execute("SELECT COUNT(*) FROM messages_libres").fetchone()[0]
        except Exception:
            msgs = "—"

        # Stats par présence
        try:
            pres_deux = conn.execute(
                "SELECT COUNT(*) FROM users WHERE presence=? AND completed=1",
                ("Oui, les deux jours",)
            ).fetchone()[0]
            pres_un   = conn.execute(
                "SELECT COUNT(*) FROM users WHERE presence=? AND completed=1",
                ("Un seul des deux jours",)
            ).fetchone()[0]
        except Exception:
            pres_deux = pres_un = "—"

        # Stats par frein
        try:
            freins = conn.execute(
                "SELECT frein, COUNT(*) as n FROM users WHERE completed=1 AND frein IS NOT NULL GROUP BY frein"
            ).fetchall()
            freins_txt = "\n".join(f"  • {r['frein']} : *{r['n']}*" for r in freins) or "  _(aucun)_"
        except Exception:
            freins_txt = "—"

        # Débutants vs traders
        try:
            debutants = conn.execute(
                "SELECT COUNT(*) FROM users WHERE deja_trade='Non' AND completed=1"
            ).fetchone()[0]
            experimentes = conn.execute(
                "SELECT COUNT(*) FROM users WHERE deja_trade='Oui' AND completed=1"
            ).fetchone()[0]
        except Exception:
            debutants = experimentes = "—"

        try:
            nb_sondages = conn.execute("SELECT COUNT(*) FROM sondages WHERE actif=1").fetchone()[0]
            nb_votes    = conn.execute("SELECT COUNT(*) FROM sondage_reponses").fetchone()[0]
        except Exception:
            nb_sondages = "—"
            nb_votes    = "—"

    await update.message.reply_text(
        f"*Statistiques Trading Pour Tous :*\n\n"
        f"👥 Membres canal : *{members}*\n\n"
        f"📝 Formulaire démarré : *{total}*\n"
        f"✅ Inscriptions complètes : *{complets}*\n"
        f"⏳ En cours : *{total - complets}*\n\n"
        f"*Présence :*\n"
        f"  • Les 2 jours : *{pres_deux}*\n"
        f"  • Un seul jour : *{pres_un}*\n\n"
        f"*Expérience :*\n"
        f"  • Débutants : *{debutants}*\n"
        f"  • Ont déjà tradé : *{experimentes}*\n\n"
        f"*Freins :*\n{freins_txt}\n\n"
        f"💬 Messages libres reçus : *{msgs}*\n\n"
        f"📊 Sondages actifs : *{nb_sondages}*\n"
        f"🗳️ Votes enregistrés : *{nb_votes}*",
        parse_mode="Markdown"
    )


# ════════════════════════════════════════════════════════════════════════════
# ── EXPORT EXCEL ──────────────────────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

async def export_users(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("Commande réservée à l'administrateur.")
        return

    await update.message.reply_text("Génération du fichier Excel en cours...")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import io

        with db() as conn:
            rows = conn.execute("""
                SELECT
                    m.telegram_id,
                    m.joined_at,
                    u.prenom,
                    u.whatsapp,
                    u.pays,
                    u.deja_trade,
                    u.interet,
                    u.presence,
                    u.frein,
                    u.level,
                    u.objectif,
                    u.email,
                    u.categorie,
                    CASE WHEN u.completed = 1 THEN 'Oui' ELSE 'Non' END AS complet,
                    u.last_seen
                FROM members_log m
                LEFT JOIN users u ON u.telegram_id = m.telegram_id
                ORDER BY m.joined_at DESC
            """).fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trading Pour Tous"

        headers = [
            "ID Telegram", "Date arrivée", "Prénom", "WhatsApp", "Pays",
            "Déjà tradé", "Intérêt principal", "Présence 2+3 sept", "Frein principal",
            "Niveau (ancien)", "Objectif (ancien)", "Email (ancien)",
            "Catégorie", "Inscription complète", "Dernière activité"
        ]
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")

        # Colonnes clés à mettre en évidence pour Charbel (présence + frein)
        highlight_fill = PatternFill("solid", fgColor="FFD966")

        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = 22

        # Colonnes 8 (présence) et 9 (frein) surlignées dans le header
        ws.cell(row=1, column=8).fill = PatternFill("solid", fgColor="C00000")
        ws.cell(row=1, column=9).fill = PatternFill("solid", fgColor="C00000")

        for row_idx, row in enumerate(rows, start=2):
            values = [
                row["telegram_id"], row["joined_at"] or "",
                row["prenom"] or "", row["whatsapp"] or "",
                row["pays"] or "", row["deja_trade"] or "",
                row["interet"] or "", row["presence"] or "",
                row["frein"] or "",
                row["level"] or "", row["objectif"] or "", row["email"] or "",
                row["categorie"] or "", row["complet"] or "Non",
                row["last_seen"] or "",
            ]
            fill = PatternFill("solid", fgColor="EBF3FB" if row_idx % 2 == 0 else "FFFFFF")
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                # Surligner présence et frein s'ils sont remplis
                if col_idx in (8, 9) and val:
                    cell.fill = highlight_fill
                else:
                    cell.fill = fill

        # ── Feuille synthèse pour Charbel : compteurs présence + frein ──
        ws2 = wb.create_sheet("Synthèse pour Charbel")
        ws2.cell(row=1, column=1, value="PRÉSENCE").font = Font(bold=True, size=14)
        ws2.cell(row=2, column=1, value="Réponse")
        ws2.cell(row=2, column=2, value="Nombre")

        pres_stats = conn.execute("""
            SELECT presence, COUNT(*) as n FROM users
            WHERE completed=1 AND presence IS NOT NULL
            GROUP BY presence
        """).fetchall()
        r = 3
        for row in pres_stats:
            ws2.cell(row=r, column=1, value=row["presence"])
            ws2.cell(row=r, column=2, value=row["n"])
            r += 1

        r += 2
        ws2.cell(row=r, column=1, value="FREINS").font = Font(bold=True, size=14)
        r += 1
        ws2.cell(row=r, column=1, value="Réponse")
        ws2.cell(row=r, column=2, value="Nombre")
        r += 1
        frein_stats = conn.execute("""
            SELECT frein, COUNT(*) as n FROM users
            WHERE completed=1 AND frein IS NOT NULL
            GROUP BY frein ORDER BY n DESC
        """).fetchall()
        for row in frein_stats:
            ws2.cell(row=r, column=1, value=row["frein"])
            ws2.cell(row=r, column=2, value=row["n"])
            r += 1

        ws2.column_dimensions["A"].width = 40
        ws2.column_dimensions["B"].width = 12

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        total     = len(rows)
        complets  = sum(1 for r in rows if r["complet"] == "Oui")
        sans_info = sum(1 for r in rows if not r["prenom"])

        await update.message.reply_document(
            document=buf,
            filename="trading_pour_tous_export.xlsx",
            caption=(
                "Export complet\n\n"
                + "Total membres : *" + str(total) + "*\n"
                + "Inscriptions complètes : *" + str(complets) + "*\n"
                + "Sans informations : *" + str(sans_info) + "*\n\n"
                + "_Onglet « Synthèse pour Charbel » : présence + freins agrégés_"
            ),
            parse_mode="Markdown"
        )

    except ImportError:
        await update.message.reply_text(
            "Le module openpyxl n'est pas installé. Lance : pip install openpyxl"
        )
    except Exception as e:
        await update.message.reply_text(f"Erreur lors de l'export : {e}")


# ════════════════════════════════════════════════════════════════════════════
# ── CONVERSATION INSCRIPTION ──────────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

async def je_me_enregistre(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    _touch_last_seen(user_id)
    if _is_already_registered(user_id):
        await _reply_already_registered(context.bot, user_id)
        return ConversationHandler.END

    await update.message.reply_text(
        "🚀 *Confirme ta place — Formation gratuite Trading Pour Tous*\n\n"
        "📅 *2 et 3 septembre, 21h00* (heure du Bénin)\n"
        "🎥 En direct uniquement — places limitées\n\n"
        "Je suis l'assistant de Charbel Yayi 👋\n"
        "Je vais te guider étape par étape pour ton inscription.\n\n"
        "*1/7 — Comment tu t'appelles ?* 😊\n\n"
        "_Réponds simplement avec ton prénom_",
        parse_mode="Markdown"
    )
    return PRENOM


async def get_prenom(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    texte   = update.message.text.strip()
    _touch_last_seen(user_id)

    prenom_candidat, confirmation_requise = _extraire_prenom(texte)

    if confirmation_requise:
        context.user_data["prenom_candidat"] = prenom_candidat
        context.user_data["prenom_original"]  = texte

        if len(prenom_candidat) > 15:
            await update.message.reply_text(
                "J'ai du mal à identifier ton prénom dans ce que tu as écrit.\n\n"
                "Peux-tu m'envoyer *uniquement ton prénom* s'il te plaît ?",
                parse_mode="Markdown"
            )
            return PRENOM

        await update.message.reply_text(
            f"Est-ce que ton prénom est *{prenom_candidat}* ?",
            parse_mode="Markdown",
            reply_markup=kb_prenom_confirm(prenom_candidat)
        )
        return PRENOM_CONFIRM

    prenom = prenom_candidat
    context.user_data["prenom"] = prenom
    upsert_user(user_id, prenom=prenom)
    return await _ask_whatsapp(update.message, prenom)


async def confirm_prenom(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query   = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    if query.data == "prenom_oui":
        prenom = context.user_data["prenom_candidat"]
        context.user_data["prenom"] = prenom
        upsert_user(user_id, prenom=prenom)
        await query.message.reply_text(f"Parfait *{prenom}* 👋", parse_mode="Markdown")
        return await _ask_whatsapp(query.message, prenom)

    await query.message.reply_text(
        "Pas de souci 😊\n\nEnvoie-moi juste ton prénom :",
        parse_mode="Markdown"
    )
    return PRENOM


async def _ask_whatsapp(message, prenom: str) -> int:
    await message.reply_text(
        f"Enchanté *{prenom}* 👋\n\n"
        "*2/7 — Quel est ton numéro WhatsApp ?*\n"
        "Je t'enverrai les rappels pour la formation 😊\n\n"
        "_(avec indicatif pays si possible, ex : +229 60619292)_",
        parse_mode="Markdown"
    )
    return WHATSAPP


async def get_whatsapp(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id  = update.message.from_user.id
    whatsapp = update.message.text.strip()
    context.user_data["whatsapp"] = whatsapp
    upsert_user(user_id, whatsapp=whatsapp)

    await update.message.reply_text(
        "*3/7 — Dans quel pays es-tu ?* 🌍\n\n"
        "_(Ex : Bénin, Côte d'Ivoire, France...)_",
        parse_mode="Markdown"
    )
    return PAYS


async def get_pays(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    pays    = update.message.text.strip()
    context.user_data["pays"] = pays
    upsert_user(user_id, pays=pays)

    await update.message.reply_text(
        "*4/7 — As-tu déjà fait du trading ?*",
        parse_mode="Markdown",
        reply_markup=kb_deja_trade()
    )
    return DEJA_TRADE


async def get_deja_trade(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query   = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    reponse = "Oui" if query.data == "trade_oui" else "Non"
    context.user_data["deja_trade"] = reponse
    upsert_user(user_id, deja_trade=reponse)

    # Logique conditionnelle : si "Non", proposer les vidéos d'initiation
    if reponse == "Non":
        await query.message.reply_text(
            "🎬 *Tu débutes en trading ?*\n\n"
            "On a préparé une série de *10 vidéos d'initiation*, spécialement pour toi, "
            "pour que tu arrives à la formation gratuite déjà à l'aise avec les bases.\n\n"
            f"👉 Clique ici pour les rejoindre : {LIEN_YOUTUBE_DEBUTANTS}\n\n"
            "_On continue ton inscription 👇_",
            parse_mode="Markdown",
            disable_web_page_preview=False
        )

    await query.message.reply_text(
        "*5/7 — Qu'est-ce qui t'intéresse le plus dans le trading ?*",
        parse_mode="Markdown",
        reply_markup=kb_interet()
    )
    return INTERET


async def get_interet(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query   = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    int_map = {
        "int_revenu":       "Un revenu complémentaire",
        "int_competence":   "Apprendre une nouvelle compétence",
        "int_independance": "L'indépendance financière",
        "int_autre":        "Autre",
    }
    interet = int_map.get(query.data, "Non précisé")
    context.user_data["interet"] = interet
    upsert_user(user_id, interet=interet)

    await query.message.reply_text(
        "*6/7 — Peux-tu être présent le 2 ET le 3 septembre à 21h ?*\n\n"
        "_⚠️ Pas d'option replay — la session est en direct uniquement_",
        parse_mode="Markdown",
        reply_markup=kb_presence()
    )
    return PRESENCE


async def get_presence(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query   = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    pres_map = {
        "pres_deux": "Oui, les deux jours",
        "pres_un":   "Un seul des deux jours",
    }
    presence = pres_map.get(query.data, "Non précisé")
    context.user_data["presence"] = presence
    upsert_user(user_id, presence=presence)

    await query.message.reply_text(
        "*7/7 — Qu'est-ce qui t'a empêché jusqu'ici de te lancer dans le trading ?*",
        parse_mode="Markdown",
        reply_markup=kb_frein()
    )
    return FREIN


async def get_frein(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query   = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    frein_map = {
        "frein_temps":      "Le manque de temps",
        "frein_peur":       "La peur de perdre de l'argent",
        "frein_commencer":  "Je ne sais pas par où commencer",
        "frein_autre":      "Autre",
    }
    frein = frein_map.get(query.data, "Non précisé")
    context.user_data["frein"] = frein
    upsert_user(user_id, frein=frein)

    prenom     = context.user_data.get("prenom", "")
    whatsapp   = context.user_data.get("whatsapp", "")
    pays       = context.user_data.get("pays", "")
    deja_trade = context.user_data.get("deja_trade", "")
    interet    = context.user_data.get("interet", "")
    presence   = context.user_data.get("presence", "")

    await query.message.reply_text(
        f"*Récapitulatif de ton inscription :*\n\n"
        f"👤 Prénom : *{prenom}*\n"
        f"📱 WhatsApp : *{whatsapp}*\n"
        f"🌍 Pays : *{pays}*\n"
        f"📊 Déjà tradé : *{deja_trade}*\n"
        f"🎯 Intérêt : *{interet}*\n"
        f"📅 Présence : *{presence}*\n"
        f"🧱 Frein : *{frein}*\n\n"
        "En confirmant, tu acceptes de recevoir les rappels par WhatsApp et Telegram.\n"
        "Tu peux te désinscrire à tout moment.\n\n"
        "👇",
        parse_mode="Markdown",
        reply_markup=kb_confirmation()
    )
    return CONFIRMATION


async def confirmer_inscription(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query   = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    prenom  = context.user_data.get("prenom", "")
    upsert_user(user_id, completed=1)
    await query.message.reply_text(
        f"🎉 *C'est noté {prenom} !*\n\n"
        "Ta place est confirmée pour la formation gratuite "
        "*Trading Pour Tous* les *2 et 3 septembre à 21h* (heure du Bénin).\n\n"
        "On se retrouve dans le canal Telegram pour tous les rappels avant le jour J.\n\n"
        "Tu recevras aussi le lien du live et les rappels par *WhatsApp et Telegram*.\n\n"
        "À très vite 🔥",
        parse_mode="Markdown"
    )
    return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Inscription annulée. Tape /JeMEnregistre pour recommencer."
    )
    return ConversationHandler.END


async def timeout_inscription(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id if update.effective_user else None
    if user_id:
        try:
            await context.bot.send_message(
                chat_id=user_id,
                text=(
                    "⏰ Ta session a expiré après 5 minutes d'inactivité.\n\n"
                    "Ton inscription n'a pas été enregistrée.\n\n"
                    "Quand tu es prêt, clique sur /JeMEnregistre pour recommencer !"
                ),
                parse_mode="Markdown"
            )
        except Exception as e:
            print(f"Timeout message uid={user_id} : {e}")


# ════════════════════════════════════════════════════════════════════════════
# ── GESTIONNAIRE D'ERREURS GLOBAL ────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE):
    tb = "".join(traceback.format_exception(
        type(context.error), context.error, context.error.__traceback__
    ))

    uname = "?"
    texte = ""
    if isinstance(update, Update):
        user  = update.effective_user
        uname = f"@{user.username}" if user and user.username else str(user.id if user else "?")
        msg   = update.message or (update.callback_query.message if update.callback_query else None)
        texte = (msg.text or "")[:100] if msg else ""

    tb_court  = tb[-2000:] if len(tb) > 2000 else tb
    ligne_sep = "\n"
    notif = (
        "*ERREUR BOT*" + ligne_sep + ligne_sep
        + "User : " + uname + ligne_sep
        + "Message : _" + texte + "_" + ligne_sep + ligne_sep
        + "`" + tb_court + "`"
    )
    print(f"[ERROR] {tb}")

    for admin_id in ADMIN_IDS:
        try:
            await context.bot.send_message(
                chat_id=admin_id,
                text=notif,
                parse_mode="Markdown"
            )
        except Exception as e:
            print(f"Impossible d'envoyer l'erreur à l'admin {admin_id} : {e}")


# ════════════════════════════════════════════════════════════════════════════
# ── MAIN ──────────────────────────────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    init_db()
    _migrate_db()   # ← IMPORTANT : décommenté pour appliquer les nouvelles colonnes
    init_sondage_db()

    app = (
        Application.builder()
        .token(TOKEN)
        .read_timeout(30)
        .write_timeout(30)
        .build()
    )

    # ── Handlers standards ────────────────────────────────────────────────
    app.add_handler(ChatJoinRequestHandler(approve_join_request))
    app.add_handler(CommandHandler("start",        start))
    app.add_handler(CommandHandler("stats",        stats))
    app.add_handler(CommandHandler("relancer",     relancer))
    app.add_handler(CommandHandler("export_users", export_users))
    app.add_handler(CallbackQueryHandler(relance_callback, pattern="^relance_go$"))

    # ── Conversations ─────────────────────────────────────────────────────
    conv_inscription = ConversationHandler(
        entry_points=[CommandHandler("JeMEnregistre", je_me_enregistre)],
        states={
            PRENOM:         [MessageHandler(filters.TEXT & ~filters.COMMAND, get_prenom)],
            PRENOM_CONFIRM: [CallbackQueryHandler(confirm_prenom, pattern="^prenom_(oui|non)$")],
            WHATSAPP:       [MessageHandler(filters.TEXT & ~filters.COMMAND, get_whatsapp)],
            PAYS:           [MessageHandler(filters.TEXT & ~filters.COMMAND, get_pays)],
            DEJA_TRADE:     [CallbackQueryHandler(get_deja_trade, pattern="^trade_(oui|non)$")],
            INTERET:        [CallbackQueryHandler(get_interet,    pattern="^int_")],
            PRESENCE:       [CallbackQueryHandler(get_presence,   pattern="^pres_")],
            FREIN:          [CallbackQueryHandler(get_frein,      pattern="^frein_")],
            CONFIRMATION:   [CallbackQueryHandler(confirmer_inscription, pattern="^confirme$")],
            ConversationHandler.TIMEOUT: [
                MessageHandler(filters.ALL, timeout_inscription),
                CallbackQueryHandler(timeout_inscription),
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_chat=False, per_user=True, allow_reentry=True,
        conversation_timeout=300,
    )

    conv_broadcast = ConversationHandler(
        entry_points=[CommandHandler("envoyer", bc_start)],
        states={
            BC_CIBLE:  [MessageHandler(filters.TEXT & ~filters.COMMAND, bc_get_cible)],
            BC_FORMAT: [MessageHandler(filters.TEXT & ~filters.COMMAND, bc_get_format)],
            BC_MEDIA:  [MessageHandler(filters.PHOTO | filters.VIDEO,   bc_get_media)],
            BC_TEXT:   [MessageHandler(filters.TEXT | filters.PHOTO | filters.VIDEO, bc_get_text)],
        },
        fallbacks=[CommandHandler("cancel", bc_cancel)],
        per_chat=False, per_user=True, allow_reentry=True,
    )

    conv_categorie = ConversationHandler(
        entry_points=[CommandHandler("nouvelle_categorie", nouvelle_categorie_start)],
        states={
            CAT_NOM: [MessageHandler(filters.TEXT & ~filters.COMMAND, nouvelle_categorie_nom)],
        },
        fallbacks=[CommandHandler("cancel", nouvelle_categorie_cancel)],
        per_chat=False, per_user=True, allow_reentry=True,
    )

    app.add_error_handler(error_handler)
    app.add_handler(conv_inscription)
    app.add_handler(conv_broadcast)
    app.add_handler(conv_categorie)

    # ── Handlers sondage ──────────────────────────────────────────────────
    register_sondage_handlers(app)

    # ── En dernier — capture tout message hors conversation ───────────────
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, message_libre))

    print("start...")
    app.run_polling(poll_interval=1, allowed_updates=Update.ALL_TYPES)